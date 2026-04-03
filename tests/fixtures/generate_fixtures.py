"""
Indian market test fixture generator for BakeManage.

Generates:
  indian_vendor_invoice.xlsx  — Excel invoice from Indian vendor (Amul / ITC style)
  indian_vendor_invoice.png   — PIL-rendered invoice image with Hindi-style vendor name
  indian_sales_receipt.xlsx   — Daily sales receipt with Indian bakery items + GST
  indian_qr_scan.png          — QR-style coloured tile image (simulates camera frame)
  indian_recipe_barfi.png     — Recipe card image (Kaju Barfi)
  indian_receipt_b2b.xlsx     — B2B purchase invoice for GSTR-1 testing

Run: python tests/fixtures/generate_fixtures.py
"""
from __future__ import annotations

import io
import os
from datetime import date, timedelta
from pathlib import Path

# ── output directory ────────────────────────────────────────────────────────
OUT = Path(__file__).parent
OUT.mkdir(parents=True, exist_ok=True)


# ── 1. Indian vendor Excel invoice ──────────────────────────────────────────
def make_excel_vendor_invoice() -> None:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Metadata columns (read by parse_excel_invoice)
    # Header row expected by BakeManage parser
    headers = [
        "vendor_name", "invoice_number", "date",
        "item_name", "quantity", "unit_price",
        "unit_of_measure", "tax_rate", "category",
        "vertical", "expiration_date",
    ]
    ws.append(headers)

    today = date.today()
    exp_flour   = (today + timedelta(days=180)).isoformat()
    exp_ghee    = (today + timedelta(days=90)).isoformat()
    exp_sugar   = (today + timedelta(days=365)).isoformat()
    exp_butter  = (today + timedelta(days=60)).isoformat()
    exp_sooji   = (today + timedelta(days=120)).isoformat()
    exp_milk    = (today + timedelta(days=7)).isoformat()
    exp_elaichi = (today + timedelta(days=365)).isoformat()
    exp_atta    = (today + timedelta(days=150)).isoformat()

    rows = [
        # vendor_name, invoice_number, date, item_name, qty, unit_price, uom, tax_rate, category, vertical, exp_date
        ["Amul Dairy Ltd", "AMU-2026-0401", today.isoformat(), "Amul Butter (500g)", 40.0, 260.0, "pcs", 12.0, "dairy_fat", "bakery", exp_butter],
        ["Amul Dairy Ltd", "AMU-2026-0401", today.isoformat(), "Amul Fresh Milk (1L)", 60.0, 68.0, "litres", 0.0, "dairy_milk", "bakery", exp_milk],
        ["ITC Foods Ltd", "ITC-2026-0402", today.isoformat(), "Aashirvaad Atta (10kg)", 20.0, 395.0, "bag", 0.0, "flour", "bakery", exp_atta],
        ["ITC Foods Ltd", "ITC-2026-0402", today.isoformat(), "Aashirvaad Maida (5kg)", 30.0, 185.0, "bag", 5.0, "flour", "bakery", exp_flour],
        ["Patanjali Ayurved", "PAT-2026-0401", today.isoformat(), "Patanjali Desi Ghee (1kg)", 25.0, 620.0, "kg", 12.0, "dairy_fat", "bakery", exp_ghee],
        ["Tate & Lyle India Pvt", "TLY-2026-0401", today.isoformat(), "Refined Sugar (50kg)", 10.0, 2350.0, "bag", 5.0, "sugar", "bakery", exp_sugar],
        ["Tate & Lyle India Pvt", "TLY-2026-0401", today.isoformat(), "Sooji (Fine Semolina) 25kg", 15.0, 850.0, "bag", 0.0, "flour", "bakery", exp_sooji],
        ["Everest Spices Ltd", "EVR-2026-0401", today.isoformat(), "Elaichi Powder (100g)", 20.0, 115.0, "pcs", 5.0, "spice", "bakery", exp_elaichi],
        ["Everest Spices Ltd", "EVR-2026-0401", today.isoformat(), "Kaju (Cashew) Grade A 1kg", 10.0, 980.0, "kg", 5.0, "dry_fruit", "bakery", exp_elaichi],
    ]

    # Style header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 16)

    for row in rows:
        ws.append(row)

    wb.save(OUT / "indian_vendor_invoice.xlsx")
    print(f"  ✓ {OUT / 'indian_vendor_invoice.xlsx'}")


# ── 2. Indian vendor invoice image (PIL) ────────────────────────────────────
def make_invoice_image() -> None:
    from PIL import Image, ImageDraw, ImageFont

    W, H = 900, 1200
    img = Image.new("RGB", (W, H), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)

    today = date.today()

    # Try to load a system font, fall back to default
    try:
        font_large  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 26)
        font_medium = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 18)
        font_small  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14)
    except (IOError, OSError):
        font_large = font_medium = font_small = ImageFont.load_default()

    # Company header block
    draw.rectangle([(0, 0), (W, 90)], fill=(31, 78, 121))
    draw.text((30, 12), "AMUL COOPERATIVE DAIRY LIMITED", fill="white", font=font_large)
    draw.text((30, 48), "GSTIN: 24AAAAC0946D1ZJ  |  Anand, Gujarat - 388001", fill=(200, 220, 255), font=font_small)
    draw.text((30, 66), "CIN: U15200GJ1946PLC000614  |  Tel: +91-2692-258506", fill=(200, 220, 255), font=font_small)

    # Invoice details
    draw.text((30, 110), f"TAX INVOICE", fill=(31, 78, 121), font=font_large)
    draw.text((30, 148), f"Invoice No: AMU-2026-{today.strftime('%m%d')}", fill=(50, 50, 50), font=font_medium)
    draw.text((30, 172), f"Invoice Date: {today.isoformat()}", fill=(50, 50, 50), font=font_medium)
    draw.text((30, 196), f"Place of Supply: Maharashtra (27)", fill=(50, 50, 50), font=font_medium)

    # Bill-To block
    draw.rectangle([(30, 230), (430, 310)], outline=(180, 180, 180), width=1)
    draw.text((40, 238), "Bill To:", fill=(31, 78, 121), font=font_medium)
    draw.text((40, 260), "BakeManage Bakery Pvt Ltd", fill=(50, 50, 50), font=font_small)
    draw.text((40, 278), "GSTIN: 27AABCB1234F1Z5", fill=(50, 50, 50), font=font_small)
    draw.text((40, 296), "Pune, Maharashtra - 411001", fill=(50, 50, 50), font=font_small)

    # Table header
    y = 340
    draw.rectangle([(30, y), (W - 30, y + 30)], fill=(230, 237, 245))
    cols = [30, 280, 380, 470, 570, 680, 780]
    headers = ["Item Description", "HSN", "Qty", "Rate (₹)", "CGST%", "SGST%", "Amount (₹)"]
    for i, h in enumerate(headers):
        draw.text((cols[i] + 5, y + 6), h, fill=(31, 78, 121), font=font_small)
    y += 30

    items = [
        ("Amul Butter (500g)", "0405", "40 pcs", "260.00", "6%", "6%", "12,416.00"),
        ("Amul Fresh Milk 1L", "0401", "60 L", "68.00", "0%", "0%", "4,080.00"),
        ("Amul Cream (200ml)", "0401", "20 pcs", "45.00", "0%", "0%", "900.00"),
        ("Amul Paneer (200g)", "0406", "15 pcs", "95.00", "6%", "6%", "1,691.70"),
        ("Condensed Milk (400g)", "1702", "30 tins", "72.00", "12%", "12%", "2,822.40"),
    ]
    for i, row in enumerate(items):
        bg = (252, 252, 252) if i % 2 == 0 else (245, 248, 252)
        draw.rectangle([(30, y), (W - 30, y + 26)], fill=bg)
        for j, cell in enumerate(row):
            draw.text((cols[j] + 5, y + 4), cell, fill=(50, 50, 50), font=font_small)
        y += 26

    # Total block
    y += 15
    draw.rectangle([(500, y), (W - 30, y + 80)], fill=(230, 237, 245))
    draw.text((510, y + 8),  "Sub Total:   ₹21,910.10", fill=(50, 50, 50), font=font_medium)
    draw.text((510, y + 30), "CGST (avg):  ₹  936.60", fill=(50, 50, 50), font=font_small)
    draw.text((510, y + 50), "SGST (avg):  ₹  936.60", fill=(50, 50, 50), font=font_small)
    draw.text((510, y + 68), "Grand Total: ₹23,783.30", fill=(31, 78, 121), font=font_large)

    # Footer
    y += 110
    draw.text((30, y), "Terms & Conditions:", fill=(31, 78, 121), font=font_medium)
    draw.text((30, y + 22), "1. Payment due within 30 days of invoice date.", fill=(100, 100, 100), font=font_small)
    draw.text((30, y + 40), "2. Goods once sold will not be taken back.", fill=(100, 100, 100), font=font_small)
    draw.text((30, y + 58), "3. Subject to Gujarat jurisdiction.", fill=(100, 100, 100), font=font_small)
    draw.text((30, y + 80), "E. & O.E.  |  This is a computer-generated invoice.", fill=(150, 150, 150), font=font_small)
    draw.text((30, H - 30), "AMUL COOPERATIVE DAIRY LIMITED  |  www.amul.com", fill=(180, 180, 180), font=font_small)

    img.save(OUT / "indian_vendor_invoice.png", format="PNG")
    print(f"  ✓ {OUT / 'indian_vendor_invoice.png'}")


# ── 3. Indian bakery sales receipt ──────────────────────────────────────────
def make_sales_receipt_excel() -> None:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Sales"

    today = date.today()

    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "BakeManage Bakery — Daily Sales Report"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Date: {today.isoformat()}"
    ws["A2"].font = Font(italic=True)

    headers = ["Product Name", "Category", "Qty Sold", "Unit Price (₹)", "GST%", "GST Amt (₹)", "Total (₹)"]
    ws.append([""] * 7)  # row 3 spacer
    ws.append(headers)
    for col in range(1, 8):
        c = ws.cell(row=4, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center")

    items = [
        ("Kaju Barfi (250g box)",     "pastries_cakes",    18, 280.0, 18, 18*280*0.18, 18*280+18*280*0.18),
        ("Gulab Jamun (10 pcs box)",  "pastries_cakes",    24, 120.0, 18, 24*120*0.18, 24*120+24*120*0.18),
        ("Sooji Halwa (500g)",         "pastries_cakes",    15, 85.0,  18, 15*85*0.18,  15*85+15*85*0.18),
        ("Jeera Biscuits (200g)",     "branded_biscuits",  40, 35.0,   5, 40*35*0.05,  40*35+40*35*0.05),
        ("Plain Rusk (400g)",         "unbranded_bread",   32, 55.0,   0, 0.0,          32*55),
        ("Elaichi Rusk (350g)",       "unbranded_bread",   28, 60.0,   0, 0.0,          28*60),
        ("Besan Ladoo (250g)",        "pastries_cakes",    20, 200.0, 18, 20*200*0.18, 20*200+20*200*0.18),
        ("Namkeen Mathri (200g)",     "branded_namkeen",   35, 45.0,  12, 35*45*0.12,  35*45+35*45*0.12),
        ("Pineapple Cake Slice",      "pastries_cakes",    50, 65.0,  18, 50*65*0.18,  50*65+50*65*0.18),
        ("Butter Croissant",          "pastries_cakes",    45, 55.0,  18, 45*55*0.18,  45*55+45*55*0.18),
    ]

    for row in items:
        ws.append([
            row[0], row[1], row[2],
            round(row[3], 2), f"{row[4]}%",
            round(row[5], 2), round(row[6], 2)
        ])

    # Totals
    ws.append([""] * 7)
    ws.append(["", "", "", "TOTAL", "",
               round(sum(r[5] for r in items), 2),
               round(sum(r[6] for r in items), 2)])
    last = ws.max_row
    ws.cell(last, 4).font = Font(bold=True)
    ws.cell(last, 6).font = Font(bold=True)
    ws.cell(last, 7).font = Font(bold=True)

    col_widths = {"A": 32, "B": 20, "C": 10, "D": 16, "E": 8, "F": 14, "G": 14}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUT / "indian_sales_receipt.xlsx")
    print(f"  ✓ {OUT / 'indian_sales_receipt.xlsx'}")


# ── 4. QR / barcode scan image ───────────────────────────────────────────────
def make_qr_image() -> None:
    from PIL import Image, ImageDraw

    # Simulate a camera frame with a QR-like tile pattern in the centre
    W, H = 640, 480
    img = Image.new("RGB", (W, H), color=(40, 40, 40))  # dark camera background
    draw = ImageDraw.Draw(img)

    # Scan frame guide
    draw.rectangle([(160, 100), (480, 380)], outline=(0, 255, 0), width=3)
    draw.text((180, 82), "Scan Table QR Code", fill=(0, 255, 0))

    # QR-like pattern (21×21 modules)
    qr_x, qr_y = 200, 130
    module = 11  # px per module
    # Finder pattern top-left
    _draw_finder(draw, qr_x, qr_y, module)
    # Finder pattern top-right
    _draw_finder(draw, qr_x + 14 * module, qr_y, module)
    # Finder pattern bottom-left
    _draw_finder(draw, qr_x, qr_y + 14 * module, module)

    # Random data modules (seed-based for reproducibility)
    import random
    rng = random.Random(42)
    for r in range(21):
        for c in range(21):
            if _in_finder(r, c):
                continue
            if rng.random() > 0.55:
                x0 = qr_x + c * module
                y0 = qr_y + r * module
                draw.rectangle([(x0, y0), (x0 + module - 1, y0 + module - 1)], fill=(0, 0, 0))

    # Table label below QR
    try:
        from PIL import ImageFont
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
    except (IOError, OSError):
        font = None
    draw.text((qr_x + 30, qr_y + 21 * module + 5), "TABLE: T-03 | BakeManage",
              fill=(255, 255, 255), font=font)

    img.save(OUT / "indian_qr_scan.png", format="PNG")
    print(f"  ✓ {OUT / 'indian_qr_scan.png'}")


def _draw_finder(draw, ox, oy, m):
    # 7×7 finder pattern
    draw.rectangle([(ox, oy), (ox + 7 * m, oy + 7 * m)], fill=(0, 0, 0))
    draw.rectangle([(ox + m, oy + m), (ox + 6 * m, oy + 6 * m)], fill=(255, 255, 255))
    draw.rectangle([(ox + 2 * m, oy + 2 * m), (ox + 5 * m, oy + 5 * m)], fill=(0, 0, 0))


def _in_finder(r, c):
    if r < 8 and c < 8:    return True
    if r < 8 and c >= 13:  return True
    if r >= 13 and c < 8:  return True
    return False


# ── 5. Recipe card image ─────────────────────────────────────────────────────
def make_recipe_image() -> None:
    from PIL import Image, ImageDraw, ImageFont

    W, H = 800, 1000
    img = Image.new("RGB", (W, H), color=(254, 249, 240))
    draw = ImageDraw.Draw(img)

    try:
        font_title  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 32)
        font_head   = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20)
        font_body   = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 16)
        font_small  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14)
    except (IOError, OSError):
        font_title = font_head = font_body = font_small = ImageFont.load_default()

    # Saffron header band
    draw.rectangle([(0, 0), (W, 80)], fill=(255, 153, 51))
    draw.text((30, 18), "BAKEMANAGE RECIPE CARD", fill="white", font=font_title)

    # Recipe name
    draw.text((30, 100), "Kaju Barfi (Cashew Fudge)", fill=(139, 69, 19), font=font_title)
    draw.text((30, 145), "Category: Indian Mithai  |  Yield: 30 pieces  |  Time: 45 min",
              fill=(100, 100, 100), font=font_small)

    # Divider
    draw.rectangle([(30, 175), (W - 30, 177)], fill=(255, 153, 51))

    # Ingredients
    draw.text((30, 192), "INGREDIENTS:", fill=(31, 78, 121), font=font_head)
    ingredients = [
        ("Kaju (Cashew)",         "250 g",   "₹245.00", "dry_fruit"),
        ("Sugar",                 "150 g",   "₹  7.50", "sugar"),
        ("Milk",                  "100 ml",  "₹  6.80", "dairy"),
        ("Ghee (Desi)",           " 10 g",   "₹  6.20", "dairy_fat"),
        ("Elaichi Powder",        "  2 g",   "₹  2.30", "spice"),
        ("Kesar (Saffron)",       "  2 strands", "₹  8.50", "spice"),
        ("Silver Varq (optional)","  1 sheet",   "₹  5.00", "garnish"),
    ]
    y = 228
    for name, qty, cost, cat in ingredients:
        draw.text((50, y), f"• {name:<28}  {qty:<12}  {cost}  [{cat}]",
                  fill=(50, 50, 50), font=font_body)
        y += 28

    draw.rectangle([(30, y + 10), (W - 30, y + 12)], fill=(200, 200, 200))
    y += 28

    # Method
    draw.text((30, y), "METHOD:", fill=(31, 78, 121), font=font_head)
    y += 32
    steps = [
        "1. Grind cashews to a fine powder using a dry grinder.",
        "2. Heat sugar and milk in a heavy pan until 1-thread consistency.",
        "3. Add cashew powder; stir continuously on low flame.",
        "4. Add ghee and elaichi powder; mix well.",
        "5. Cook until mixture leaves sides of pan (approx 8-10 min).",
        "6. Transfer to greased plate; roll to 6mm thickness.",
        "7. Optionally top with silver varq; let cool 20 minutes.",
        "8. Cut into diamond shapes; store at room temp up to 5 days.",
    ]
    for step in steps:
        draw.text((50, y), step, fill=(60, 60, 60), font=font_body)
        y += 28

    # GST note
    y += 15
    draw.rectangle([(30, y), (W - 30, y + 50)], fill=(230, 245, 230))
    draw.text((40, y + 8),  "GST Note: Kaju Barfi falls under HSN 1704 — GST @ 18% (CGST 9% + SGST 9%)",
              fill=(0, 100, 0), font=font_small)
    draw.text((40, y + 26), "FSSAI Category: Dairy-Based Confection  |  Shelf-life: 5 days at 25°C",
              fill=(0, 100, 0), font=font_small)

    # Footer
    draw.rectangle([(0, H - 40), (W, H)], fill=(255, 153, 51))
    draw.text((30, H - 28), "BakeManage © 2026  |  Recipe v1.0  |  Not for commercial redistribution",
              fill="white", font=font_small)

    img.save(OUT / "indian_recipe_barfi.png", format="PNG")
    print(f"  ✓ {OUT / 'indian_recipe_barfi.png'}")


# ── 6. B2B purchase invoice for GSTR-1 testing ──────────────────────────────
def make_b2b_invoice_excel() -> None:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B2B Invoice"

    today = date.today()

    # Same columns as indian_vendor_invoice.xlsx for parse_excel_invoice compatibility
    headers = [
        "vendor_name", "invoice_number", "date",
        "item_name", "quantity", "unit_price",
        "unit_of_measure", "tax_rate", "category",
        "vertical", "expiration_date",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")

    exp_sugar = (today + timedelta(days=365)).isoformat()
    exp_flour = (today + timedelta(days=180)).isoformat()

    # B2B purchase — supplies for large catering order
    rows = [
        ["Hindustan Unilever Ltd", f"HUL-{today.strftime('%Y%m%d')}-001", today.isoformat(),
         "Dalda Vanaspati Ghee 15kg", 5.0, 1850.0, "tin", 12.0, "dairy_fat", "bakery", exp_sugar],
        ["Hindustan Unilever Ltd", f"HUL-{today.strftime('%Y%m%d')}-001", today.isoformat(),
         "Kissan Mixed Fruit Jam 1kg", 12.0, 180.0, "jar", 12.0, "flavouring", "bakery", (today + timedelta(days=365)).isoformat()],
        ["RSGSM (Rajasthan)", f"RSG-{today.strftime('%Y%m%d')}-004", today.isoformat(),
         "Chakki Fresh Atta 25kg", 8.0, 760.0, "bag", 0.0, "flour", "bakery", exp_flour],
        ["RSGSM (Rajasthan)", f"RSG-{today.strftime('%Y%m%d')}-004", today.isoformat(),
         "Besan (Chickpea Flour) 10kg", 6.0, 640.0, "bag", 5.0, "flour", "bakery", exp_flour],
        ["Parle Products Pvt Ltd", f"PAR-{today.strftime('%Y%m%d')}-009", today.isoformat(),
         "Parle-G Biscuits (Bulk 10kg)", 10.0, 950.0, "carton", 5.0, "branded_biscuits", "bakery",
         (today + timedelta(days=90)).isoformat()],
    ]

    for r in rows:
        ws.append(r)

    for col in ws.columns:
        max_w = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(max_w, 30)

    wb.save(OUT / "indian_receipt_b2b.xlsx")
    print(f"  ✓ {OUT / 'indian_receipt_b2b.xlsx'}")


# ── 7. Small browning test images ────────────────────────────────────────────
def make_browning_images() -> None:
    from PIL import Image
    import numpy as np

    # "Good" bake — golden brown (pixel intensity ~140-165 = browning 55-65%)
    arr_good = np.full((100, 100, 3), (155, 120, 70), dtype="uint8")
    Image.fromarray(arr_good).save(OUT / "bread_correctly_baked.png")
    print(f"  ✓ {OUT / 'bread_correctly_baked.png'}")

    # "Over-baked" — dark brown (pixel intensity ~60 = browning ~24%)
    arr_over = np.full((100, 100, 3), (60, 40, 20), dtype="uint8")
    Image.fromarray(arr_over).save(OUT / "bread_overbaked.png")
    print(f"  ✓ {OUT / 'bread_overbaked.png'}")

    # "Under-baked" — pale (pixel intensity ~220 = browning ~86%)
    arr_under = np.full((100, 100, 3), (220, 210, 190), dtype="uint8")
    Image.fromarray(arr_under).save(OUT / "bread_underbaked.png")
    print(f"  ✓ {OUT / 'bread_underbaked.png'}")


# ── main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Generating BakeManage Indian market test fixtures...")
    make_excel_vendor_invoice()
    make_invoice_image()
    make_sales_receipt_excel()
    make_qr_image()
    make_recipe_image()
    make_b2b_invoice_excel()
    make_browning_images()
    print("Done.")
